import openpyxl
from pathlib import Path

work3 = Path(r'D:\OneDrive\自学编程\claude code\work3')
xlsx_files = list(work3.glob('*.xlsx'))
for f in xlsx_files:
    print(f"File: {f.name}")
    wb = openpyxl.load_workbook(str(f))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}, rows={ws.max_row}, cols={ws.max_column}")
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                val = cell.value
                if val is not None:
                    row_data.append(repr(val))
            if row_data:
                print('  ', row_data)
