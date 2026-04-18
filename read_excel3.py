import openpyxl
from pathlib import Path
import json

work3 = Path(r'D:\OneDrive\自学编程\claude code\work3')
xlsx_files = list(work3.glob('*.xlsx'))
results = []

for f in xlsx_files:
    print(f"Processing: {f.name}")
    wb = openpyxl.load_workbook(str(f))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            if any(v is not None for v in row_data):
                results.append(row_data)
                print('  Row:', row_data)

# Write to output.json
out_path = work3 / 'output.json'
with open(out_path, 'w', encoding='utf-8') as fp:
    json.dump(results, fp, ensure_ascii=False, indent=2)
print('\nSaved to output.json')
